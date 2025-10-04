import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import faiss
import re
import os
import json
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass
import google.generativeai as genai
from sklearn.cluster import DBSCAN
import logging
from pathlib import Path
import time
from decouple import config
import structlog
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
import numpy as np

# Configure logging
logger = structlog.get_logger()

@dataclass
class CellContext:
    """Data class to store cell context information"""
    sheet_name: str
    table_title: str
    row_headers: List[str]
    col_headers: List[str]
    value: Any
    data_type: str
    cell_address: str
    full_context: str
    confidence_score: float = 0.0

@dataclass
class ValidationResult:
    """Data class for validation results"""
    query_value: str
    query_context: str
    matches: List[Dict[str, Any]]
    best_match_confidence: float
    validation_status: str

@dataclass
class AnalysisResult:
    """Data transfer object for analysis results"""
    faiss_index: Any
    faiss_index_path: Any
    analysedValues: Any

class ExcelAuditSystem:
    """
    Comprehensive Excel-PDF Audit System using FAISS and Gemini embeddings
    """

    def __init__(self):
        """
        Initialize the audit system

        Args:
            gemini_api_key: Google Gemini API key for embeddings
        """
        # Get API key from environment
        self.context_database: List[CellContext] = []
        self.faiss_index = None
        self.embeddings = []

        api_key = config('GOOGLE_API_KEY', default=None)
        if not api_key:
            raise ValueError("GOOGLE_API_KEY is required for Gemini 2.5 Pro")
        
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel('gemini-2.0-flash-exp')
        self.ai_enabled = True

        logger.info("Enhanced Gemini 2.5 Pro EXCEL Service initialized with comprehensive extraction settings")


        # Data type patterns (fixed regex with raw strings)
        self.data_patterns = {
            'percentage': r'^[\d.,]+%$',
            'currency': r'^[\$₹€£][\d,.]+ *[KMB]?$',
            'ratio': r'^[\d.]+:[\d.]+$|^[\d.]+x$',
            'decimal': r'^[\d.,]+$',
            'growth': r'^[+-]?[\d.,]+%?$',
            'integer': r'^[\d,]+$'
        }

    def process_sheet(self, excel_file_path: str, sheet_name: str) -> Tuple[str, List[CellContext], int]:
        """Process one sheet: detect tables, extract contexts"""
        workbook = load_workbook(excel_file_path, data_only=True)
        worksheet = workbook[sheet_name]

        tables = self._detect_tables(worksheet)
        contexts = []
        for table in tables:
            contexts.extend(self._extract_context(worksheet, table, sheet_name))

        return (sheet_name, contexts, len(tables))

    def analyse_excel_comprehensive(self, excel_file_path: str) -> AnalysisResult:
        """
        Main entry point for comprehensive Excel analysis
        """
        logger.info(f"Starting comprehensive analysis of: {excel_file_path}")

        try:
            self.context_database.clear()
            self.embeddings.clear()
            self.faiss_index = None

            workbook = load_workbook(excel_file_path, data_only=True)
            sheetnames = workbook.sheetnames
            del workbook  # free memory since each process loads workbook separately

            analysis_results = {
                'file_path': excel_file_path,
                'sheets_processed': 0,
                'tables_detected': 0,
                'contexts_extracted': 0,
                'embeddings_created': 0,
                'processing_status': 'success'
            }

            # 🔹 Parallel across sheets
            from concurrent.futures import ProcessPoolExecutor, as_completed
            with ProcessPoolExecutor() as executor:
                futures = {executor.submit(self.process_sheet, excel_file_path, sheet): sheet for sheet in sheetnames}

                for f in as_completed(futures):
                    sheet_name, contexts, tables_count = f.result()
                    self.context_database.extend(contexts)
                    analysis_results['contexts_extracted'] += len(contexts)
                    analysis_results['tables_detected'] += tables_count
                    analysis_results['sheets_processed'] += 1
                    logger.info(f"Finished sheet: {sheet_name} | Contexts: {len(contexts)}")

            # 🔹 Embedding creation (parallelized)
            if self.context_database:
                self._create_embeddings()
                analysis_results['embeddings_created'] = len(self.embeddings)
                faiss_index = self._build_faiss_index()
            else:
                faiss_index = None

            # Save contexts JSON
            context_json_path = "faiss_db/contexts.json"
            os.makedirs(os.path.dirname(context_json_path), exist_ok=True)
            with open(context_json_path, "w", encoding="utf-8") as f:
                json.dump([c.__dict__ for c in self.context_database], f, ensure_ascii=False, indent=4, default=str)
            logger.info(f"Contexts saved to {context_json_path}")

            return AnalysisResult(
                faiss_index=faiss_index,
                faiss_index_path="faiss_db/faiss_index.index",
                analysedValues=len(self.context_database)
            )

        except Exception as e:
            logger.error(f"Error in comprehensive analysis: {str(e)}")
            raise  # 🔑 don’t return dict, let caller catch

    def _detect_tables(self, worksheet) -> List[Dict[str, Any]]:
        """Multi-stage table detection"""
        tables = []

        # Get all non-empty cells
        non_empty_cells = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    non_empty_cells.append((cell.row, cell.column, cell.value))

        if not non_empty_cells:
            return tables

        # Stage 1: Find data regions using clustering
        coordinates = np.array([(row, col) for row, col, _ in non_empty_cells])

        if len(coordinates) > 5:
            try:
                clustering = DBSCAN(eps=3, min_samples=3).fit(coordinates)
                labels = clustering.labels_

                # Group cells by cluster
                clusters = {}
                for i, label in enumerate(labels):
                    if label != -1:  # Ignore noise points
                        if label not in clusters:
                            clusters[label] = []
                        clusters[label].append(non_empty_cells[i])

                # Stage 2: Extract table boundaries for each cluster
                for cluster_id, cells in clusters.items():
                    if len(cells) >= 6:  # Minimum size for a table
                        table = self._extract_table_boundaries(cells)
                        if table:
                            # Stage 3: Classify table type
                            table['type'] = self._classify_table_type(cells)
                            tables.append(table)

            except Exception as e:
                logger.warning(f"Clustering failed, using fallback method: {str(e)}")
                # Fallback: treat all non-empty cells as one table
                table = self._extract_table_boundaries(non_empty_cells)
                if table:
                    table['type'] = 'mixed'
                    tables.append(table)

        return tables

    def _extract_table_boundaries(self, cells: List[Tuple]) -> Optional[Dict[str, Any]]:
        """Extract table boundaries from cell list"""
        if not cells:
            return None

        rows = [cell[0] for cell in cells]
        cols = [cell[1] for cell in cells]

        return {
            'min_row': min(rows),
            'max_row': max(rows),
            'min_col': min(cols),
            'max_col': max(cols),
            'cells': cells,
            'size': len(cells)
        }

    def _classify_table_type(self, cells: List[Tuple]) -> str:
        """Classify table type based on content patterns"""
        values = [str(cell[2]).lower() for cell in cells if cell[2] is not None]

        # Check for financial indicators
        financial_keywords = ['ratio', 'growth', 'mix', 'profit', 'revenue', 'pbt']
        percentage_count = sum(1 for v in values if '%' in v)

        if any(keyword in ' '.join(values) for keyword in financial_keywords):
            if percentage_count > len(values) * 0.3:
                return 'financial_ratios'
            else:
                return 'financial_summary'
        elif percentage_count > len(values) * 0.5:
            return 'percentage_table'
        else:
            return 'data_table'

    def _extract_context(self, worksheet, table: Dict[str, Any], sheet_name: str) -> List[CellContext]:
        """Hierarchical context building for each cell"""
        contexts = []

        # Find table title
        table_title = self._get_table_title(worksheet, table)

        # Process each cell in the table
        for row_idx in range(table['min_row'], table['max_row'] + 1):
            for col_idx in range(table['min_col'], table['max_col'] + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                if cell.value is not None:
                    # Get hierarchical headers
                    row_headers = self._get_row_hierarchy(worksheet, table, row_idx)
                    col_headers = self._get_col_hierarchy(worksheet, table, col_idx)

                    # Classify data type
                    data_type = self._classify_data_types(cell.value)

                    # Build full context string
                    full_context = self._build_context_string(
                        sheet_name, table_title, row_headers, col_headers, 
                        cell.value, data_type, table['type']
                    )

                    context = CellContext(
                        sheet_name=sheet_name,
                        table_title=table_title,
                        row_headers=row_headers,
                        col_headers=col_headers,
                        value=cell.value,
                        data_type=data_type,
                        cell_address=f"{cell.column_letter}{cell.row}",
                        full_context=full_context
                    )

                    contexts.append(context)

        return contexts

    def _get_table_title(self, worksheet, table: Dict[str, Any]) -> str:
        """Extract table title from above the table region"""
        # Look 1-3 rows above the table for title
        for row_offset in range(1, 4):
            title_row = table['min_row'] - row_offset
            if title_row > 0:
                for col_idx in range(table['min_col'], table['max_col'] + 1):
                    cell = worksheet.cell(row=title_row, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        title = str(cell.value).strip()
                        if len(title) > 3 and not title.replace('.', '').isdigit():
                            return title

        return f"Table_{table['min_row']}_{table['min_col']}"

    def _get_row_hierarchy(self, worksheet, table: Dict[str, Any], row_idx: int) -> List[str]:
        """Get hierarchical row headers"""
        headers = []

        # Check leftmost columns for row headers
        for col_offset in range(min(3, table['max_col'] - table['min_col'] + 1)):
            col_idx = table['min_col'] + col_offset
            cell = worksheet.cell(row=row_idx, column=col_idx)

            if cell.value and isinstance(cell.value, str):
                header = str(cell.value).strip()
                if header and not self._is_numeric_value(header):
                    headers.append(header)
                    break

        return headers if headers else ['Unknown_Row']

    def _get_col_hierarchy(self, worksheet, table: Dict[str, Any], col_idx: int) -> List[str]:
        """Get hierarchical column headers"""
        headers = []

        # Check top rows for column headers
        for row_offset in range(min(3, table['max_row'] - table['min_row'] + 1)):
            row_idx = table['min_row'] + row_offset
            cell = worksheet.cell(row=row_idx, column=col_idx)

            if cell.value and isinstance(cell.value, str):
                header = str(cell.value).strip()
                if header and not self._is_numeric_value(header):
                    headers.append(header)
                    break

        return headers if headers else ['Unknown_Col']

    def _classify_data_types(self, value: Any) -> str:
        """Handle percentages, ratios, amounts classification"""
        if value is None:
            return 'empty'

        value_str = str(value).strip()

        # Check each pattern (using raw strings to fix regex)
        for data_type, pattern in self.data_patterns.items():
            if re.match(pattern, value_str, re.IGNORECASE):
                return data_type

        # Check for text
        if isinstance(value, str) and not value_str.replace(' ', '').replace('.', '').replace(',', '').isdigit():
            return 'text'

        return 'unknown'

    def _is_numeric_value(self, value_str: str) -> bool:
        """Check if string represents a numeric value"""
        for pattern in self.data_patterns.values():
            if re.match(pattern, value_str, re.IGNORECASE):
                return True
        return False

    def _build_context_string(self, sheet_name: str, table_title: str, 
                             row_headers: List[str], col_headers: List[str], 
                             value: Any, data_type: str, table_type: str) -> str:
        """Build comprehensive context string for embedding"""
        context_parts = [
            f"Sheet: {sheet_name}",
            f"Table: {table_title}",
            f"Type: {table_type}",
            f"Row: {' > '.join(row_headers)}",
            f"Column: {' > '.join(col_headers)}",
            f"Value: {value}",
            f"DataType: {data_type}"
        ]

        return " | ".join(context_parts)

    def _create_embeddings(self) -> None:
        """Generate Gemini embeddings for all contexts using threading"""
        logger.info("Creating embeddings for contexts...")

        texts = [c.full_context for c in self.context_database]
        self.embeddings = []

        def embed_text(text):
            try:
                result = genai.embed_content(model="models/embedding-001", content=text)
                return result['embedding']
            except Exception as e:
                logger.error(f"Embedding failed: {e}")
                return np.random.rand(768).tolist()  # fallback random vec

        with ThreadPoolExecutor(max_workers=10) as executor:  # tune workers based on API rate limit
            futures = {executor.submit(embed_text, t): t for t in texts}
            for i, f in enumerate(as_completed(futures)):
                emb = f.result()
                self.embeddings.append(emb)
                if i % 50 == 0:
                    logger.info(f"Generated {i} embeddings...")

        logger.info(f"Generated total {len(self.embeddings)} embeddings")

    def _build_faiss_index(self) -> None:
        """Optimize for similarity search using FAISS"""
        if not self.embeddings:
            raise ValueError("No embeddings available to build index")

        # Convert embeddings to numpy array
        embedding_matrix = np.array(self.embeddings).astype('float32')

        # Create FAISS index
        dimension = embedding_matrix.shape[1]
        faiss_index = faiss.IndexFlatIP(dimension)

        # Normalize vectors for cosine similarity
        faiss.normalize_L2(embedding_matrix)

        # Add embeddings to index
        faiss_index.add(embedding_matrix)

        logger.info(f"FAISS index built with {faiss_index.ntotal} vectors")
        save_path = "faiss_db/faiss_index.index"
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        
        faiss.write_index(faiss_index, save_path)
        logger.info(f"FAISS index saved to {save_path}")

        return faiss_index

# Usage Example:
def main():
    """Example usage of the ExcelAuditSystem"""
    # Initialize system
    audit_system = ExcelAuditSystem()

    # Analyze Excel file
    excel_path = "/Users/himanshusharma/Personal_Code/sample data/Slide 1.xlsx"
    results = audit_system.analyse_excel_comprehensive(excel_path)
    print("Analysis Results:", results)
    

if __name__ == "__main__":
    main()

