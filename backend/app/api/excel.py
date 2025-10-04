# app/api/excel.py
from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel
from typing import List, Any, Tuple
from openpyxl import load_workbook
import os
import re

router = APIRouter()

# ---- Resolve Excel path for a given file_id --------------------
def resolve_excel_path(file_id: str) -> str:
    """
    Resolve file_id to actual Excel file path.
    Simple file-based lookup in uploads directory.
    """
    from pathlib import Path
    
    # Get the uploads directory path
    # Go up from app/api/excel.py -> app -> backend -> uploads
    backend_root = Path(__file__).parent.parent.parent
    uploads_dir = backend_root / "uploads"
    
    # The file_id IS the filename (UUID from your upload system)
    file_path = uploads_dir / file_id
    
    # Check if file exists
    if file_path.exists() and file_path.is_file():
        return str(file_path)
    
    # Try with common Excel extensions
    for ext in ['.xlsx', '.xls', '.xlsm']:
        file_with_ext = uploads_dir / f"{file_id}{ext}"
        if file_with_ext.exists() and file_with_ext.is_file():
            return str(file_with_ext)
    
    # File not found
    raise HTTPException(
        status_code=404, 
        detail=f"Excel file not found in uploads: {file_id}"
    )

# ---- Utilities --------------------------------------------------
_A1_PATTERN = re.compile(r"^([A-Za-z]+)(\d+)$")

def a1_to_rc(a1: str) -> Tuple[int, int]:
    """Convert A1 notation (e.g., 'B5') to (row, col) 1-based indices"""
    m = _A1_PATTERN.match(a1.strip())
    if not m:
        raise HTTPException(400, detail=f"Invalid A1 notation: {a1}")
    
    letters, row = m.groups()
    col = 0
    for ch in letters.upper():
        col = col * 26 + (ord(ch) - 64)
    
    return int(row), col

# ---- Schemas ----------------------------------------------------
class SheetMeta(BaseModel):
    name: str
    rows: int
    cols: int

class MetaResponse(BaseModel):
    file_id: str
    sheets: List[SheetMeta]

class PageResponse(BaseModel):
    sheet: str
    r0: int   # 1-based inclusive
    r1: int   # 1-based inclusive
    c0: int   # 1-based inclusive
    c1: int   # 1-based inclusive
    data: List[List[Any]]

class SpotlightResponse(BaseModel):
    sheet: str
    row: int
    col: int

# ---- Endpoints --------------------------------------------------

@router.get("/meta", response_model=MetaResponse)
def excel_meta(file_id: str):
    """
    Get sheet list and dimensions for an Excel file.
    GET /api/excel/meta?file_id=xxx
    """
    path = resolve_excel_path(file_id)
    
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets.append(SheetMeta(
                name=sheet_name,
                rows=ws.max_row or 1,
                cols=ws.max_column or 1
            ))
        
        wb.close()
        return MetaResponse(file_id=file_id, sheets=sheets)
    
    except Exception as e:
        raise HTTPException(400, detail=f"Workbook open error: {str(e)}")


@router.get("/page", response_model=PageResponse)
def excel_page(
    file_id: str,
    sheet: str,
    r0: int = Query(..., ge=1, description="Start row (1-based)"),
    r1: int = Query(..., ge=1, description="End row (1-based, inclusive)"),
    c0: int = Query(..., ge=1, description="Start column (1-based)"),
    c1: int = Query(..., ge=1, description="End column (1-based, inclusive)"),
):
    """
    Return a tile of Excel data (e.g., 200 rows × 50 cols).
    GET /api/excel/page?file_id=xxx&sheet=Sheet1&r0=1&r1=200&c0=1&c1=50
    """
    if r1 < r0 or c1 < c0:
        raise HTTPException(400, detail="Invalid range: end must be >= start")
    
    path = resolve_excel_path(file_id)
    
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        
        if sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(404, detail=f"Sheet not found: {sheet}")
        
        ws = wb[sheet]
        max_r = ws.max_row or 1
        max_c = ws.max_column or 1
        
        # Clamp to actual bounds
        r0_clamped = max(1, min(r0, max_r))
        r1_clamped = max(1, min(r1, max_r))
        c0_clamped = max(1, min(c0, max_c))
        c1_clamped = max(1, min(c1, max_c))
        
        # Fetch data using iter_rows for efficiency
        rows = ws.iter_rows(
            min_row=r0_clamped,
            max_row=r1_clamped,
            min_col=c0_clamped,
            max_col=c1_clamped,
            values_only=True
        )
        
        # Convert to list, replacing None with empty string
        data = [[("" if v is None else v) for v in row] for row in rows]
        
        wb.close()
        
        return PageResponse(
            sheet=sheet,
            r0=r0_clamped,
            r1=r1_clamped,
            c0=c0_clamped,
            c1=c1_clamped,
            data=data
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, detail=f"Read error: {str(e)}")


@router.get("/spotlight", response_model=SpotlightResponse)
def excel_spotlight(file_id: str, sheet: str, cell: str):
    """
    Quick helper to get row/col for a specific cell (for centering).
    GET /api/excel/spotlight?file_id=xxx&sheet=Sheet1&cell=B5
    """
    path = resolve_excel_path(file_id)
    
    try:
        r, c = a1_to_rc(cell)
        
        wb = load_workbook(path, read_only=True, data_only=True)
        
        if sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(404, detail=f"Sheet not found: {sheet}")
        
        ws = wb[sheet]
        
        # Clamp to bounds
        r = max(1, min(r, ws.max_row or 1))
        c = max(1, min(c, ws.max_column or 1))
        
        wb.close()
        
        return SpotlightResponse(sheet=sheet, row=r, col=c)
    
    except Exception as e:
        raise HTTPException(400, detail=f"Spotlight error: {str(e)}")